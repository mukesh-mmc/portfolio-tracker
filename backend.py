# ==============================
# Portfolio Tracker - App Ready Backend
# Refactored from v2.2
# ==============================

import logging
import shutil
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# -------- LOGGING --------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# -------- NAV CACHE --------
_nav_history_cache = {}
_nav_latest_cache = {}

# ==============================
# NAV FUNCTIONS
# ============================== 

def _fetch_nav_history(scheme_code, retries=3):
    if scheme_code in _nav_history_cache:
        log.info(f"[CACHE HIT] Scheme {scheme_code}: Returning cached NAV history")
        return _nav_history_cache[scheme_code]

    url = f"https://api.mfapi.in/mf/{scheme_code}"
    log.info(f"[API CALL] Fetching NAV history for Scheme {scheme_code} from {url}")

    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=5)
            r.raise_for_status()
            data = r.json()

            if "data" not in data:
                log.warning(f"[API ERROR] Scheme {scheme_code}: 'data' key not found in API response")
                return None

            log.info(f"[API SUCCESS] Scheme {scheme_code}: Fetched {len(data['data'])} records. Latest NAV: {data['data'][0]['nav']} on {data['data'][0]['date']}")
            _nav_history_cache[scheme_code] = data["data"]
            time.sleep(0.3)
            return data["data"]

        except Exception as e:
            log.warning(f"[API RETRY] Scheme {scheme_code}, Attempt {attempt + 1}/{retries}: {str(e)}")
            continue

    log.error(f"[API FAILED] Scheme {scheme_code}: Failed after {retries} retries")
    return None


def get_nav_data(scheme_code):
    if scheme_code in _nav_latest_cache:
        cached_result = _nav_latest_cache[scheme_code]
        log.info(f"[CACHE HIT] Scheme {scheme_code}: Using cached NAV {cached_result[0]} from {cached_result[2]}")
        return cached_result

    log.info(f"[CACHE MISS] Scheme {scheme_code}: Cache miss, fetching from API")
    history = _fetch_nav_history(scheme_code)
    if not history:
        log.error(f"[NAV ERROR] Scheme {scheme_code}: No history data available")
        return None, None, None

    latest = float(history[0]["nav"])
    prev = float(history[1]["nav"]) if len(history) > 1 else latest
    nav_date = pd.to_datetime(history[0]["date"], dayfirst=True).date()

    log.info(f"[NAV FETCHED] Scheme {scheme_code}: Latest NAV {latest}, Previous NAV {prev}, Date {nav_date}")

    _nav_latest_cache[scheme_code] = (latest, prev, nav_date)
    return latest, prev, nav_date


def get_nav_by_date(scheme_code, target_date):
    log.info(f"[NAV BY DATE] Scheme {scheme_code}: Looking for NAV on or before {target_date}")
    history = _fetch_nav_history(scheme_code)
    if not history:
        log.error(f"[NAV BY DATE ERROR] Scheme {scheme_code}: No history data available")
        return None

    target_dt = pd.to_datetime(target_date).date()

    for row in history:
        nav_date = pd.to_datetime(row["date"], dayfirst=True).date()
        if nav_date <= target_dt:
            nav_value = float(row["nav"])
            log.info(f"[NAV BY DATE SUCCESS] Scheme {scheme_code}: Found NAV {nav_value} on {nav_date}")
            return nav_value

    log.warning(f"[NAV BY DATE NOT FOUND] Scheme {scheme_code}: No NAV found for date {target_date}")
    return None

# ==============================
# FINANCIAL CALCULATIONS
# ==============================

def calculate_xirr(cashflows, guess=0.1):
    cashflows = [(pd.to_datetime(d).date(), cf) for d, cf in cashflows]
    base_date = cashflows[0][0]

    def xnpv(rate):
        return sum(cf / (1 + rate) ** ((d - base_date).days / 365) for d, cf in cashflows)

    rate = guess

    for _ in range(100):
        f_val = xnpv(rate)
        d_rate = 1e-6
        f_deriv = (xnpv(rate + d_rate) - f_val) / d_rate

        if f_deriv == 0:
            return None

        new_rate = rate - f_val / f_deriv
        if abs(new_rate - rate) < 1e-6:
            return new_rate

        rate = new_rate

    return None


def adjust_to_working_day(date):
    if date.weekday() == 5:
        return date + pd.Timedelta(days=2)
    elif date.weekday() == 6:
        return date + pd.Timedelta(days=1)
    return date

# ==============================
# DATA LOADING
# ==============================

def load_transactions(file_path):
    df = pd.read_excel(file_path, sheet_name="Transactions", engine="openpyxl")

    df.columns = df.columns.str.strip()
    df["Transaction Date"] = pd.to_datetime(df["Transaction Date"])
    df["Units"] = pd.to_numeric(df["Units"], errors="coerce")
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    df["Scheme Code"] = pd.to_numeric(df["Scheme Code"], errors="coerce")

    df = df.dropna(subset=["Units", "Amount", "Scheme Code"])
    df["Scheme Code"] = df["Scheme Code"].astype(int)

    return df


def load_sip(file_path):
    try:
        sip_df = pd.read_excel(file_path, sheet_name="SIP", engine="openpyxl")
        sip_df.columns = sip_df.columns.str.strip()
        return sip_df
    except Exception:
        return None

# ==============================
# SIP PROCESSING
# ============================== 

def process_sip(df, sip_df):
    if sip_df is None:
        return df

    today = pd.Timestamp.today().date()
    new_rows = []

    for _, sip in sip_df.iterrows():
        scheme = sip["Scheme Name"]
        code = int(sip["Scheme Code"])
        sip_day = int(sip["Day"])
        amount = float(sip["Amount"])

        for offset in [0, -1]:
            base_date = pd.Timestamp.today() + pd.DateOffset(months=offset)
            raw_date = base_date.replace(day=min(sip_day, 28)).date()
            sip_date = adjust_to_working_day(raw_date)

            if sip_date > today:
                continue

            already_exists = (
                (df["Scheme Name"] == scheme) &
                (df["Transaction Date"].dt.date == sip_date)
            ).any()

            if already_exists:
                continue

            nav = get_nav_by_date(code, sip_date)
            if nav is None:
                continue

            new_rows.append({
                "Transaction Date": sip_date,
                "Scheme Name": scheme,
                "Units": round(amount / nav, 3),
                "Amount": amount,
                "Scheme Code": code,
                "Price": round(nav, 2),
            })

    if new_rows:
        df = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)

    return df

# ==============================
# PORTFOLIO CALCULATION
# ============================== 

def calculate_portfolio(df):
    rows = []
    today = pd.Timestamp.today().date()

    for scheme, group in df.groupby("Scheme Name"):
        units = group["Units"].sum()
        if units == 0:
            continue

        code = int(group["Scheme Code"].iloc[0])
        net_inv = group["Amount"].sum()

        latest, prev, nav_date = get_nav_data(code)
        if latest is None:
            continue

        curr_val = units * latest
        prev_val = units * prev
        daily = curr_val - prev_val
        daily_pct = (daily / prev_val * 100) if prev_val else 0
        total_ret = curr_val - net_inv

        cashflows = [(r["Transaction Date"], -r["Amount"]) for _, r in group.iterrows()]
        cashflows.append((pd.Timestamp.today(), curr_val))

        xirr = calculate_xirr(cashflows)

        rows.append({
            "Scheme Name": scheme,
            "Scheme Code": code,
            "Units": round(units, 3),
            "Net Investment (₹)": round(net_inv, 2),
            "Latest NAV (₹)": round(latest, 3),
            "Previous NAV (₹)": round(prev, 3),
            "Daily Change (₹)": round(daily, 2),
            "Daily Change (%)": round(daily_pct, 2),
            "Current Value (₹)": round(curr_val, 2),
            "Total Return (₹)": round(total_ret, 2),
            "XIRR (%)": round(xirr * 100, 2) if xirr else None,
            "As of Date": nav_date,
        })

        portfolio_df = pd.DataFrame(rows)

    # -------- TOTAL ROW --------
    if not portfolio_df.empty:
        total_investment = portfolio_df["Net Investment (₹)"].sum()
        current_value = portfolio_df["Current Value (₹)"].sum()
        total_return = portfolio_df["Total Return (₹)"].sum()
        daily_total = portfolio_df["Daily Change (₹)"].sum()

        prev_total_val = current_value - daily_total
        daily_pct_total = (daily_total / prev_total_val * 100) if prev_total_val else 0

        portfolio_cashflows = [
            (row["Transaction Date"], -row["Amount"]) for _, row in df.iterrows()
        ]
        portfolio_cashflows.append((pd.Timestamp.today(), current_value))

        portfolio_xirr = calculate_xirr(portfolio_cashflows)

        total_row = pd.DataFrame([{
            "Scheme Name": "TOTAL",
            "Scheme Code": "",
            "Units": "",
            "Net Investment (₹)": round(total_investment, 2),
            "Latest NAV (₹)": "",
            "Previous NAV (₹)": "",
            "Daily Change (₹)": round(daily_total, 2),
            "Daily Change (%)": round(daily_pct_total, 2),
            "Current Value (₹)": round(current_value, 2),
            "Total Return (₹)": round(total_return, 2),
            "XIRR (%)": round(portfolio_xirr * 100, 2) if portfolio_xirr else None,
            "As of Date": "",
        }])

        portfolio_df = pd.concat([portfolio_df, total_row], ignore_index=True)

    return portfolio_df

# ==============================
# EXCEL FORMATTING
# ============================== 

def format_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb["Portfolio"]

    green_font = Font(color="008000")
    red_font = Font(color="FF0000")
    bold_font = Font(bold=True)

    # Map column names
    col_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    currency_cols = [
        "Net Investment (₹)", "Latest NAV (₹)", "Previous NAV (₹)",
        "Daily Change (₹)", "Current Value (₹)", "Total Return (₹)"
    ]

    percent_cols = ["Daily Change (%)", "XIRR (%)"]

    color_cols = ["Daily Change (₹)", "Daily Change (%)", "Total Return (₹)"]

    for row in ws.iter_rows(min_row=2):
        for col_name in currency_cols:
            col_num = col_index.get(col_name)
            if col_num:
                cell = row[col_num - 1]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "₹#,##0.00"

        for col_name in percent_cols:
            col_num = col_index.get(col_name)
            if col_num:
                cell = row[col_num - 1]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

        for col_name in color_cols:
            col_num = col_index.get(col_name)
            if col_num:
                cell = row[col_num - 1]
                if isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.font = green_font
                    elif cell.value < 0:
                        cell.font = red_font

        # Bold TOTAL row
        if row[0].value == "TOTAL":
            for cell in row:
                cell.font = bold_font

    wb.save(file_path)

# ==============================
# MAIN RUN FUNCTION (APP ENTRY)
# ============================== 

def run_portfolio(file_path, save_output=True):
    """
    Main function to be used by Streamlit / API
    Returns portfolio dataframe
    """
    log.info(f"[RUN START] Portfolio calculation started at {datetime.now()}")

    file_path = Path(file_path)

    # Backup
    if file_path.exists():
        backup = file_path.with_suffix(f".{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak.xlsx")
        shutil.copy2(file_path, backup)
        log.info(f"[BACKUP] Created backup at {backup}")

    # Load
    df = load_transactions(file_path)
    sip_df = load_sip(file_path)
    log.info(f"[LOAD] Loaded {len(df)} transactions")

    # Process SIP
    df = process_sip(df, sip_df)
    log.info(f"[SIP PROCESS] After SIP processing: {len(df)} total transactions")

    # Calculate
    portfolio_df = calculate_portfolio(df)
    log.info(f"[CALCULATE] Portfolio calculated with {len(portfolio_df)} rows")

    # Save
    if save_output:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            portfolio_df.to_excel(writer, sheet_name="Portfolio", index=False)
            df.to_excel(writer, sheet_name="Transactions", index=False)

        format_excel(file_path)
        log.info(f"[SAVE] Portfolio saved and formatted")

    log.info(f"[RUN END] Portfolio calculation completed at {datetime.now()}")
    return portfolio_df
